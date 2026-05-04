
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_test_plan():
    """
    Generates a detailed test plan Excel file for the SehatVault project.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet

    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    category_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # --- Test Cases Data ---
    test_cases = {
        "Backend API": [
            # Authentication
            ("AUTH-01", "Authentication", "Successful Patient Signup", "Verify a new patient can register successfully with valid, unique details.", "1. Send a POST request to /api/auth/signup with a unique email, strong password, and valid hospital ID.", "HTTP 201 Created. User is created in the database with 'PATIENT' role. A success message is returned.", "", "", ""),
            ("AUTH-02", "Authentication", "Successful Admin Signup", "Verify a new admin can register successfully with valid, unique details.", "1. Send a POST request to /api/auth/signup/admin with a unique email and strong password.", "HTTP 201 Created. User is created with 'ADMIN' role.", "", "", ""),
            ("AUTH-03", "Authentication", "Signup with Duplicate Email", "Verify the system prevents registration with an email that already exists.", "1. Use the email from AUTH-01.\n2. Send another POST request to /api/auth/signup with the same email.", "HTTP 409 Conflict or 400 Bad Request. Error message indicating the email is already in use.", "", "", ""),
            ("AUTH-04", "Authentication", "Successful Signin", "Verify a registered user can log in with correct credentials.", "1. Use credentials from AUTH-01.\n2. Send a POST request to /api/auth/signin.", "HTTP 200 OK. A valid JWT token and user details are returned.", "", "", ""),
            ("AUTH-05", "Authentication", "Signin with Invalid Password", "Verify login fails with an incorrect password.", "1. Use email from AUTH-01 but a wrong password.\n2. Send a POST request to /api/auth/signin.", "HTTP 401 Unauthorized. Error message indicating invalid credentials.", "", "", ""),
            # KYC & Profile
            ("KYC-01", "KYC & Profile", "Submit KYC Information", "A user with a PENDING KYC status can submit their profile information for verification.", "1. Log in as a patient (from AUTH-01).\n2. Send a POST request to the /api/kyc/submit endpoint with all required fields (name, address, ID documents).", "HTTP 200 OK. The user's KYC status changes to 'SUBMITTED'. The data is saved correctly in the database.", "", "", ""),
            ("KYC-02", "KYC & Profile", "Admin Views KYC Submissions", "Verify an admin can retrieve a list of all KYC submissions pending approval.", "1. Log in as an admin.\n2. Send a GET request to /api/kyc/submissions.", "HTTP 200 OK. A list of users with 'SUBMITTED' KYC status is returned.", "", "", ""),
            ("KYC-03", "KYC & Profile", "Admin Approves KYC", "Verify an admin can approve a KYC submission.", "1. Log in as an admin.\n2. Send a PATCH request to /api/kyc/approve/{userId} for a user with 'SUBMITTED' status.", "HTTP 200 OK. The user's KYC status changes to 'VERIFIED'.", "", "", ""),
            ("KYC-04", "KYC & Profile", "Admin Rejects KYC", "Verify an admin can reject a KYC submission.", "1. Log in as an admin.\n2. Send a PATCH request to /api/kyc/reject/{userId} with a reason for rejection.", "HTTP 200 OK. The user's KYC status changes to 'REJECTED'.", "", "", ""),
            # Asset Deposits
            ("DEPOSIT-01", "Asset Deposit", "Create Deposit Request (KYC Verified)", "A KYC-verified patient can successfully request to deposit an asset.", "1. Log in as a 'VERIFIED' patient (from KYC-03).\n2. Send a POST request to /api/deposits with asset details (type, value, description).", "HTTP 201 Created. A new deposit record is created with 'PENDING' status.", "", "", ""),
            ("DEPOSIT-02", "Asset Deposit", "Create Deposit Request (KYC Not Verified)", "A patient who is not KYC-verified cannot request a deposit.", "1. Log in as a 'PENDING' patient.\n2. Send a POST request to /api/deposits.", "HTTP 403 Forbidden. An error message indicates KYC is required.", "", "", ""),
            ("DEPOSIT-03", "Asset Deposit", "Admin Approves Deposit", "Verify an admin can approve a pending asset deposit.", "1. Log in as an admin.\n2. Send a PATCH request to /api/deposits/approve/{depositId}.", "HTTP 200 OK. The deposit status changes to 'APPROVED'. This should trigger the blockchain service to mint AT.", "", "", ""),
        ],
        "Frontend UI-UX": [
            # Dashboard
            ("DASH-01", "Dashboard", "Patient Dashboard Loads Correctly", "Verify the patient dashboard displays correct summary data after login.", "1. Log in as a patient with existing assets and balances.\n2. Navigate to the dashboard URL.", "The dashboard loads without errors. Total AT, Available AT, In Trade AT, and HT Balance are all visible and accurate.", "", "", ""),
            ("DASH-02", "Dashboard", "KYC Pending Banner Interaction", "Verify the 'Update Profile' link on the KYC banner navigates to the correct page.", "1. Log in as a new user with 'PENDING' KYC.\n2. Click the 'Update Profile' or 'Open KYC' link in the yellow banner.", "The user is redirected to the profile or KYC submission page.", "", "", ""),
            ("DASH-03", "Dashboard", "Refresh Button Functionality", "Verify the 'Refresh' button updates the dashboard data.", "1. Log in as a patient.\n2. Have an admin perform an action in the background (e.g., approve a deposit).\n3. Click the 'Refresh' button on the dashboard.", "The dashboard data updates to reflect the new state (e.g., Approved Deposits count increases).", "", "", ""),
            # Asset Deposit Flow
            ("DEPOSIT-F-01", "Asset Deposit", "Navigate to Deposit Form", "Verify user can navigate to the 'Deposit Asset' page from the sidebar.", "1. Log in as a patient.\n2. Click on 'Deposit Asset' in the left navigation menu.", "The 'Deposit Asset' form page is displayed.", "", "", ""),
            ("DEPOSIT-F-02", "Asset Deposit", "Form Validation (Empty)", "Verify the deposit form shows validation errors if submitted empty.", "1. Navigate to the 'Deposit Asset' page.\n2. Click the 'Submit' button without filling any fields.", "Validation messages appear under each required field (e.g., 'Asset type is required').", "", "", ""),
            ("DEPOSIT-F-03", "Asset Deposit", "Successful Form Submission", "Verify a success message is shown after a valid deposit request is submitted.", "1. Fill out the deposit form with valid data.\n2. Click 'Submit'.", "A success toast/notification appears (e.g., 'Your deposit request has been submitted!'). The user is redirected to the activity page.", "", "", ""),
            # Marketplace
            ("MARKET-01", "Marketplace", "View Available Assets", "Verify the marketplace correctly displays assets available for trading.", "1. Log in as a patient.\n2. Navigate to the 'Marketplace' page.", "A list or grid of assets, potentially from other users, is displayed with key details like value and potential return.", "", "", ""),
            # Notifications
            ("NOTIF-01", "Notifications", "View Notifications Panel", "Verify that clicking the notification bell opens a panel with a list of notifications.", "1. Log in as a user with unread notifications.\n2. Click the bell icon in the header.", "A dropdown or sidebar appears, listing recent notifications with timestamps.", "", "", ""),
            ("NOTIF-02", "Notifications", "Mark as Read", "Verify that viewing notifications clears the 'new' indicator.", "1. Perform steps from NOTIF-01.\n2. Close the notification panel.", "The red dot/counter on the bell icon disappears.", "", "", ""),
        ],
        "Blockchain (Smart Contracts)": [
            # Use Case 1: Asset Tokenization
            ("BC-UC1-01", "UC1: Asset Tokenization", "Admin Mints AssetToken", "Verify an authorized admin can mint AssetTokens (AT) to a patient after an asset is approved.", "1. Use a Hardhat script.\n2. Call `mintAssetToken` on HospitalFinancials with a patient address, unique depositId, and amount.", "The patient's AT balance on the AssetToken contract increases. The `AssetTokenMinted` event is emitted with correct details.", "", "", ""),
            ("BC-UC1-02", "UC1: Asset Tokenization", "Non-Admin Fails to Mint", "Verify a random account cannot mint AssetTokens.", "1. Connect to the contract with an account that does not have the MINTER_ROLE.\n2. Attempt to call `mint` directly on the AssetToken contract.", "The transaction reverts with an 'AccessControl' error message.", "", "", ""),
            ("BC-UC1-03", "UC1: Asset Tokenization", "Mint with Zero Amount", "Verify minting zero tokens does not change the state.", "1. Call `mintAssetToken` with an amount of 0.\n2. Check the patient's balance before and after.", "The transaction may succeed but the patient's AT balance should remain unchanged. No event should be emitted or it should emit with amount 0.", "", "", ""),
            # Use Case 2: Investment & Profit
            ("BC-UC2-01", "UC2: Investment & Profit", "Record a Profitable Trade", "Verify admin can record a successful trade and its profit.", "1. Call `recordTrade` on HospitalFinancials with `investedAT` > 0 and `profit` > 0.", "A new Trade struct is created with the correct values and a new `tradeId`. The `TradeRecorded` event is emitted.", "", "", ""),
            ("BC-UC2-02", "UC2: Investment & Profit", "Record a Trade with Loss", "Verify a trade can be recorded with zero or negative profit (if logic allows).", "1. Call `recordTrade` with `investedAT` > 0 and `profit` = 0.", "The trade is recorded successfully. This tests the system's ability to handle break-even or loss scenarios.", "", "", ""),
            # Use Case 3: Profit Distribution
            ("BC-UC3-01", "UC3: Profit Distribution", "Distribute HealthTokens", "Verify admin can distribute profits as HealthTokens (HT) to multiple recipients.", "1. After a trade is recorded (BC-UC2-01), call `distributeProfit` with the `tradeId` and matching arrays of recipients and amounts.", "Each recipient's HT balance on the HealthToken contract increases. The `ProfitDistributed` event is emitted.", "", "", ""),
            ("BC-UC3-02", "UC3: Profit Distribution", "Distribution Mismatch", "Verify transaction reverts if recipient and amount arrays have different lengths.", "1. Call `distributeProfit` with `recipients` array of length 2 and `amountsHT` array of length 3.", "The transaction reverts with the error 'Arrays length mismatch'.", "", "", ""),
            ("BC-UC3-03", "UC3: Profit Distribution", "Distribute from Invalid Trade", "Verify transaction reverts if `tradeId` does not exist.", "1. Call `distributeProfit` with a `tradeId` that has not been recorded (e.g., 999).", "The transaction reverts with an error indicating the trade does not exist.", "", "", ""),
            # Use Case 4: Benefit Redemption
            ("BC-UC4-01", "UC4: Benefit Redemption", "Redeem HealthToken for Service", "Verify admin can redeem (burn) a patient's HT for a healthcare service.", "1. Ensure a patient has an HT balance (from BC-UC3-01).\n2. Call `redeemHealthToken` with the patient's address and an amount less than or equal to their balance.", "The patient's HT balance decreases. The `HealthTokenRedeemed` event is emitted.", "", "", ""),
            ("BC-UC4-02", "UC4: Benefit Redemption", "Redeem with Insufficient Funds", "Verify a patient cannot redeem more HT than they own.", "1. Check a patient's HT balance.\n2. Call `redeemHealthToken` with an amount greater than their balance.", "The transaction reverts with an 'ERC20: burn amount exceeds balance' error.", "", "", ""),
        ],
        "Non-Functional": [
            # Performance
            ("PERF-01", "Performance", "API Response Time (High Load)", "Ensure all critical API endpoints respond within an acceptable time frame under high load.", "1. Use JMeter to send 200 concurrent requests to /api/dashboard/patient/summary.\n2. Measure the 95th percentile response time.", "95th percentile response time should be less than 2000ms.", "", "", ""),
            ("PERF-02", "Performance", "Database Query Efficiency", "Verify that pages with large amounts of data (e.g., activity log) do not cause slow database queries.", "1. Seed the database with 10,000 activity records for a single user.\n2. Load the activity page for that user and measure server response time.", "The page should load in under 3 seconds. Check query logs for unindexed queries.", "", "", ""),
            # Security
            ("SEC-01", "Security", "Unauthorized Endpoint Access", "Verify that unauthenticated users cannot access any protected API endpoints.", "1. Without a JWT token, send requests to /api/profile, /api/deposits, /api/wallet.\n2. Observe the HTTP status code.", "All requests should be rejected with a 401 Unauthorized or 403 Forbidden status.", "", "", ""),
            ("SEC-02", "Security", "Cross-User Data Access (Horizontal)", "Verify a logged-in patient cannot access another patient's data.", "1. Log in as Patient A.\n2. Attempt to send a GET request to an endpoint using Patient B's ID (e.g., /api/wallet/{patient_B_id}).", "The request should be rejected with 403 Forbidden or return no data.", "", "", ""),
            ("SEC-03", "Security", "Role-Based Access Control (Vertical)", "Verify a patient cannot access admin-only endpoints.", "1. Log in as a patient.\n2. Attempt to send a PATCH request to an admin endpoint like /api/kyc/approve/{userId}.", "The request should be rejected with 403 Forbidden.", "", "", ""),
            ("SEC-04", "Security", "Input Sanitization (XSS)", "Verify user inputs are sanitized to prevent Cross-Site Scripting (XSS).", "1. In a form field (e.g., profile name), enter a script tag like `<script>alert('XSS')</script>`.\n2. Submit the form and view the data elsewhere in the app.", "The script should not execute. The output should be displayed as plain text (e.g., `&lt;script&gt;...`).", "", "", ""),
            # Usability & Accessibility
            ("USAB-01", "Usability", "Responsive Design (Mobile/Tablet)", "Ensure the application is fully usable and readable on various screen sizes.", "1. Open the frontend in a browser.\n2. Use developer tools to simulate different devices (iPhone, iPad, Android).\n3. Navigate through all major user flows.", "The layout adjusts gracefully. All buttons are clickable, text is readable, and no horizontal scrolling is required.", "", "", ""),
            ("USAB-02", "Accessibility", "Keyboard Navigation", "Verify all interactive elements can be accessed and used via keyboard.", "1. Using only the Tab key, navigate through the login page and dashboard.\n2. Ensure all buttons, links, and form fields can be focused.\n3. Use Enter/Space to activate elements.", "A visible focus indicator is present. All interactive elements are reachable and functional in a logical order.", "", "", ""),
        ]
    }

    # --- Create Sheets and Headers ---
    for sheet_name, tests in test_cases.items():
        sheet = workbook.create_sheet(title=sheet_name)
        headers = ["Test Case ID", "Category", "Feature", "Test Case Description", "Steps to Test", "Expected Result", "Actual Result", "Status", "Notes"]
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # --- Populate Data ---
        for row_num, test_data in enumerate(tests, 2):
            # Apply category fill to the entire row for visual grouping
            if row_num > 2 and tests[row_num-3][1] != test_data[1]:
                 for col_num in range(1, len(headers) + 1):
                    sheet.cell(row=row_num, column=col_num).fill = category_fill

            for col_num, value in enumerate(test_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = left_alignment
                cell.border = thin_border

        # --- Add Data Validation for Status column ---
        dv = DataValidation(type="list", formula1='"Pass,Fail,Not Tested"', allow_blank=True)
        dv.error = 'Your entry is not in the list.'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select from the list.'
        dv.promptTitle = 'Select Status'
        sheet.add_data_validation(dv)
        dv.add(f'H2:H{len(tests)+1}') # Apply to all status cells

        # --- Adjust Column Widths ---
        sheet.column_dimensions['A'].width = 15  # Test Case ID
        sheet.column_dimensions['B'].width = 20  # Category
        sheet.column_dimensions['C'].width = 25  # Feature
        sheet.column_dimensions['D'].width = 40  # Test Case Description
        sheet.column_dimensions['E'].width = 40  # Steps to Test
        sheet.column_dimensions['F'].width = 40  # Expected Result
        sheet.column_dimensions['G'].width = 40  # Actual Result
        sheet.column_dimensions['H'].width = 15  # Status
        sheet.column_dimensions['I'].width = 40  # Notes

    # --- Save the workbook ---
    output_path = "D:/Projects/FYP/SehatVault_Test_Plan.xlsx"
    try:
        workbook.save(output_path)
        print(f"Successfully created test plan: {output_path}")
    except Exception as e:
        print(f"Error saving file: {e}")

if __name__ == "__main__":
    create_test_plan()
